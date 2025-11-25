"""
Products Export/Import routes - Excel operations
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import json
from datetime import datetime

from database import get_db, Product
from auth import require_admin

router = APIRouter()

@router.get("/api/admin/products/export")
async def export_products(
    db: Session = Depends(get_db),
    current_user = Depends(require_admin)
):
    """Export all products to Excel file"""
    
    # Get all products
    products = db.query(Product).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Define headers
    headers = [
        "id", "name", "collection", "price", "image", "images",
        "description", "features", "specs", "in_stock", "stock_quantity",
        "sku", "is_featured", "movement", "case_material", "dial_color",
        "water_resistance", "seo_title", "seo_description", "seo_keywords",
        "fb_title", "fb_description", "created_at", "updated_at"
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1).value = product.id
        ws.cell(row=row_num, column=2).value = product.name
        ws.cell(row=row_num, column=3).value = product.collection
        ws.cell(row=row_num, column=4).value = product.price
        ws.cell(row=row_num, column=5).value = product.image
        ws.cell(row=row_num, column=6).value = product.images  # JSON string
        ws.cell(row=row_num, column=7).value = product.description
        ws.cell(row=row_num, column=8).value = product.features  # JSON string
        ws.cell(row=row_num, column=9).value = product.specs  # JSON string
        ws.cell(row=row_num, column=10).value = product.in_stock
        ws.cell(row=row_num, column=11).value = product.stock_quantity
        ws.cell(row=row_num, column=12).value = product.sku
        ws.cell(row=row_num, column=13).value = product.is_featured
        ws.cell(row=row_num, column=14).value = product.movement
        ws.cell(row=row_num, column=15).value = product.case_material
        ws.cell(row=row_num, column=16).value = product.dial_color
        ws.cell(row=row_num, column=17).value = product.water_resistance
        ws.cell(row=row_num, column=18).value = product.seo_title
        ws.cell(row=row_num, column=19).value = product.seo_description
        ws.cell(row=row_num, column=20).value = product.seo_keywords
        ws.cell(row=row_num, column=21).value = product.fb_title
        ws.cell(row=row_num, column=22).value = product.fb_description
        ws.cell(row=row_num, column=23).value = product.created_at.isoformat() if product.created_at else ""
        ws.cell(row=row_num, column=24).value = product.updated_at.isoformat() if product.updated_at else ""
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    filename = f"orient_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/api/admin/products/import")
async def import_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(require_admin)
):
    """Import products from Excel file"""
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")
    
    try:
        # Read file
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Validate required headers
        required_headers = ["name", "collection", "price"]
        for required in required_headers:
            if required not in headers:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required column: {required}"
                )
        
        # Process rows
        created_count = 0
        updated_count = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # Create dict from row
                row_data = dict(zip(headers, row))
                
                # Skip empty rows
                if not row_data.get("name"):
                    continue
                
                # Check if product exists by SKU or ID
                existing_product = None
                if row_data.get("sku"):
                    existing_product = db.query(Product).filter(Product.sku == row_data["sku"]).first()
                if not existing_product and row_data.get("id"):
                    existing_product = db.query(Product).filter(Product.id == row_data["id"]).first()
                
                # Prepare data
                product_data = {
                    "name": row_data["name"],
                    "collection": row_data["collection"],
                    "price": float(row_data["price"]) if row_data.get("price") else 0,
                    "image": row_data.get("image"),
                    "images": row_data.get("images"),  # Keep as string
                    "description": row_data.get("description"),
                    "features": row_data.get("features"),  # Keep as string
                    "specs": row_data.get("specs"),  # Keep as string
                    "in_stock": bool(row_data.get("in_stock", True)),
                    "stock_quantity": int(row_data.get("stock_quantity", 0)) if row_data.get("stock_quantity") else 0,
                    "sku": row_data.get("sku"),
                    "is_featured": bool(row_data.get("is_featured", False)),
                    "movement": row_data.get("movement"),
                    "case_material": row_data.get("case_material"),
                    "dial_color": row_data.get("dial_color"),
                    "water_resistance": row_data.get("water_resistance"),
                    "seo_title": row_data.get("seo_title"),
                    "seo_description": row_data.get("seo_description"),
                    "seo_keywords": row_data.get("seo_keywords"),
                    "fb_title": row_data.get("fb_title"),
                    "fb_description": row_data.get("fb_description"),
                }
                
                if existing_product:
                    # Update existing product
                    for key, value in product_data.items():
                        if value is not None:
                            setattr(existing_product, key, value)
                    existing_product.updated_at = datetime.utcnow()
                    updated_count += 1
                else:
                    # Create new product
                    product_id = row_data.get("id")
                    if not product_id:
                        # Generate ID from name
                        product_id = row_data["name"].lower().replace(" ", "-").replace("&", "and")
                        
                        # Check if ID exists
                        existing_id = db.query(Product).filter(Product.id == product_id).first()
                        if existing_id:
                            import uuid
                            product_id = f"{product_id}-{str(uuid.uuid4())[:8]}"
                    
                    new_product = Product(id=product_id, **product_data)
                    db.add(new_product)
                    created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                continue
        
        # Commit changes
        db.commit()
        
        return {
            "success": True,
            "created": created_count,
            "updated": updated_count,
            "errors": errors,
            "message": f"Импорт завершен: создано {created_count}, обновлено {updated_count}"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")